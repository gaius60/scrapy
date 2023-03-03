import selenium
import openpyxl
from selenium.webdriver import Firefox
from selenium.webdriver.firefox.options import Options
import time
import os
from PIL import Image
import datetime

# Open the Excel file
workbook = openpyxl.load_workbook('Tunisia_2023-02-23_2023-03-01.xlsx')

# Select the worksheet
worksheet = workbook['Sheet']

# Create a new directory to store the screenshots
if not os.path.exists('media'):
    os.makedirs('media')

# Configure Firefox options
options = Options()
options.add_argument('-headless')

# Initialize the web driver
driver = Firefox(options=options)

# Initialize a dictionary to keep track of previously used names
name_dict = {}

# Loop through each row in column B and take a screenshot of the page
for row in range(2, worksheet.max_row + 1):
    url = worksheet.cell(row, 2).value
    date_time = worksheet.cell(row, 4).value
    if url is not None:
        try:
            # Navigate to the URL
            driver.get(url)
            # Wait for the page to load
          
            # Take a screenshot of the page
            datetime_str = datetime.datetime.strftime(date_time, "%Y-%m-%d %H:%M:%S")
            if datetime_str in name_dict:
                name_dict[datetime_str] += 1
                screenshot_name = f"{datetime_str}.{name_dict[datetime_str]}.png"
            else:
                name_dict[datetime_str] = 0
                screenshot_name = f"{datetime_str}.png"
            screenshot_path = os.path.join('media', screenshot_name)
            driver.save_screenshot(screenshot_path)
        except selenium.common.exceptions.WebDriverException:
            print(f"Error: Could not navigate to {url}")
            continue
        
# Close the web driver
driver.quit()

print('Screenshots saved successfully!')

