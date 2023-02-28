# in coop with chatgpt
import feedparser
import openpyxl
import matplotlib.pyplot as plt
from datetime import datetime, timedelta
import matplotlib.dates as mdates
import urllib.parse

def google_news_search(query, start_date, end_date):
    query_encoded = urllib.parse.quote(query)
    base_url = 'https://news.google.com/rss/search'
    language_codes = ['en-US', 'en-GB', 'fr-FR', 'ar-TN']  # add the language codes you want to aggregate for
    location_codes = ['US', 'GB', 'FR', 'TN']  # add the location codes you want to aggregate for
    feeds = []
    titles = set()

    for lang_code in language_codes:
        for loc_code in location_codes:
            url = f'{base_url}?q={query_encoded}&hl={lang_code}&gl={loc_code}&ceid={lang_code}:{loc_code}&since={start_date}&until={end_date}'
            feed = feedparser.parse(url)
            feeds.append(feed)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(['Title', 'Link', 'Website', 'Publication Date', 'Language', 'Location'])

    start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
    end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')

    for feed in feeds:
        for entry in feed.entries:
            pub_date = datetime.strptime(entry.published, '%a, %d %b %Y %H:%M:%S %Z')
            if start_date_obj <= pub_date <= end_date_obj:
                title = entry.title
                if title not in titles:
                    link = entry.link
                    website = entry.source.title
                    language = feed.feed.language
                    location = feed.feed.title.split(' - ')[-1]
                    ws.append([title, link, website, pub_date, language, location])
                    titles.add(title)

    filename = f'{query}_{start_date}_{end_date}.xlsx'
    wb.save(filename)
    print(f'Results saved to {filename}')

    # Open the Excel file and select the correct worksheet
    workbook = openpyxl.load_workbook(filename)
    worksheet = workbook.active

    # Initialize variables
    date_count = {}

    # Loop through the rows and extract the publication dates and counts
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        # Check if the row contains the expected format
        if not row[3]:
            continue

        # Extract the publication date and count
        date = row[3]
        date_str = date.strftime('%d %b %Y')

        # Group the articles by publication date and time
        if date_str in date_count:
            date_count[date_str] += 1
        else:
            date_count[date_str] = 1

    # Sort the dates list in ascending order
    dates = sorted(date_count.keys(), key=lambda x: datetime.strptime(x, '%d %b %Y'))

    print("Ordered dates:")
    for date in dates:
        print(date)

    # Count the number of articles published every day
    counts = [date_count[date] for date in dates]

    # Print the counts
    print("Counts:")
    print(counts)

    # Set the start and end date for plotting
    start_date = datetime.strptime(min(dates), '%d %b %Y')
    end_date = datetime.strptime(max(dates), '%d %b %Y')

    # Convert datetime objects to matplotlib dates
    mdates_dates = [mdates.date2num(datetime.strptime(date, '%d %b %Y')) for date in dates]

    # Generate the graph
    plt.plot(mdates_dates, counts, color='red', marker='o', label='Number of Articles')
    plt.xlabel('Publication Date')
    plt.ylabel('Number of Articles')
    plt.title('Number of Articles Published Over Time')
    plt.gca().xaxis.set_major_formatter(mdates.DateFormatter('%d %b %Y'))
    plt.gcf().autofmt_xdate()
    plt.legend()
    plt.show()

query = input("Enter your query: ")
start_date = input("Enter start date (YYYY-MM-DD): ")
end_date = input("Enter end date (YYYY-MM-DD): ")

google_news_search(query, start_date, end_date)

